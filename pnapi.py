import requests
import openpyxl

api_key = "2b10VV3OGk2MmoZn3fnhr9p7e"
base_api_url = "https://my-api.plantnet.org/"
project = "identificacao-botanica"
def get_species_info(species_name):
    endpoint = "species/"
    params = {"name": species_name, "organs": "flower,fruit,leaf"}
    headers = {"Api-Key": api_key}
    response = requests.get(base_api_url + endpoint, headers=headers, params=params)
    if response.status_code == 200:
        results = response.json().get("results")
        if results:
            species_data = results[0]
            common_name = species_data.get("common_names")[0]
            habit = species_data.get("habitat")
            family = species_data.get("family").get("name")
            group = species_data.get("group").get("name")
            return common_name, habit, family, group
    else:
        print("erro")

# Define as espécies para pesquisar
species_list = ["Ginkgo biloba", "Lilium longiflorum", "Prunus avium"]

# Cria um arquivo Excel e adiciona as informações das espécies
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Nome da espécie", "Nome comum", "Hábito", "Grupo botânico", "Família"])



for species_name in species_list:
    species_info = get_species_info(species_name)
    if species_info:
            # Escreve os dados da espécie de planta na linha seguinte
        ws["A2"] = species_info["vernacularNames"][0]["name"] if species_info["vernacularNames"] else ""
        ws["B2"] = species_info["habit"][0]["name"] if species_info["habit"] else ""
        ws["C2"] = species_info["group"][0]["name"] if species_info["group"] else ""
        ws["D2"] = species_info["family"]["name"] if species_info["family"] else ""

wb.save("species_info.xlsx")